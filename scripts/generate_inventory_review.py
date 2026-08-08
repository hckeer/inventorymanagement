#!/usr/bin/env python3
"""Generate a client-review workbook from the provisional inventory seed.

This creates a review artefact only.  It does not apply migrations, write to
Supabase, or create print-ready barcode labels.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SEED = ROOT / "supabase/migrations/20260807184000_seed_inventory.sql"
OUTPUT = ROOT / "docs/generated/inventory_barcode_client_review.xlsx"

# These are deliberately human-facing names, not manufacturer model codes.
# The client-review workbook exposes every generated value for approval.
COMMON_NAMES = {
    "CL-40MKB": "CSTAND40",
    "CL-20MKB": "CSTAND20",
    "CL-30MKB": "CSTAND30",
    "CL-40MKB-STAND": "CSTAND40",
    "CL-20MKB-STAND": "CSTAND20",
    "CL-30MKB-STAND": "CSTAND30",
    "CL-40MKB-STAND": "CSTANDBASE40",
    "CL-20MKB-STAND": "CSTANDBASE20",
    "CL-30MKB-STAND": "CSTANDBASE30",
    "L1.0033520": "ORBITER",
    "L1.0048488": "SKYPANELX",
    "MS60R-2K-KIT": "MS60R2K",
    "TP2R-K4": "TP2RK4",
    "TP4R-K4": "TP4RK4",
    "MG1200Bi": "MG1200",
    "QUAD DOLLY": "QUADDOLLY",
    "KCP-200": "GRIPHEAD",
    "L2.0033540": "ORBITEROPTIC15",
    "L2.0033541": "ORBITEROPTIC30",
    "L2.0033542": "ORBITEROPTIC60",
    "F800R": "FLEXMAT800",
    "F100R": "FLEXMAT100",
}

KIT_COMMON_NAMES = {
    "CL-40MKB": "CSTAND40",
    "CL-20MKB": "CSTAND20",
    "CL-30MKB": "CSTAND30",
    "L1.0033520": "ORBITER",
    "L1.0048488": "SKYPANELX",
    "MS60R-2K-KIT": "MS60R2K",
    "TP2R-K4": "TP2RK4",
    "TP4R-K4": "TP4RK4",
}

# Explicit PDF rows that are required quantity components for the two ARRI kits.
ARRI_BULK_COMPONENTS = {
    "L1.0033520": ["L2.0033732", "L2.0033800", "L2.0001486", "L2.0034624"],
    "L1.0048488": ["L2.0048844", "L2.0033799", "L2.0007516"],
}

# C-stand heads and arms are shared bulk hardware, never serialized children.
# The source seed used provisional per-kit asset rows for these; this review
# converts those rows into generic quantity requirements instead.
CSTAND_BULK_COMPONENTS = {
    "CL-40MKB-HEAD": ("KCP-200", '2-1/2" Grip Head', "GRIPHEAD"),
    "CL-30MKB-HEAD": ("KCP-200", '2-1/2" Grip Head', "GRIPHEAD"),
    "CL-20MKB-HEAD": ("KCP-200B", '2-1/2" Grip Head (Black)', "GRIPHEAD20"),
    "CL-40MKB-ARM": ("KCP-241", '40" Grip Arm', "GRIPARM40"),
    "CL-30MKB-ARM": ("KCP-241", '40" Grip Arm', "GRIPARM40"),
    "CL-20MKB-ARM": ("KCP-221B", '20" Grip Arm (Black)', "GRIPARM20"),
}


def sql_rows(text: str, table: str, fields: list[str]) -> list[dict[str, str]]:
    field_pattern = r"\s*,\s*".join(re.escape(field) for field in fields)
    pattern = re.compile(
        rf"insert into public\.{re.escape(table)} \({field_pattern}\)\s+values \((.+?)\)\s+on conflict",
        re.MULTILINE | re.DOTALL,
    )
    rows = []
    for match in pattern.finditer(text):
        values = [
            (quoted.replace("''", "'") if quoted else bare)
            for quoted, bare in re.findall(r"'((?:''|[^'])*)'|([^,\s]+)", match.group(1))
        ]
        if len(values) != len(fields):
            raise ValueError(f"Could not parse {table} row: {match.group(0)}")
        rows.append(dict(zip(fields, values)))
    return rows


def common_name(product: dict[str, str]) -> str:
    if product["sku"] in COMMON_NAMES:
        return COMMON_NAMES[product["sku"]]
    name = product["name"].upper()
    name = re.sub(r"\b(KUPO|GODOX|KNOWLED|ARRI|NSH|BLACK|SILVER|WITH|WITHOUT|W/|FOR|AND|THE)\b", " ", name)
    tokens = re.findall(r"[A-Z0-9]+", name)
    # A short, readable candidate—not a hidden manufacturer-code fallback.
    return "".join(tokens[:3])[:20] or "REVIEWNAME"


def detailed_common_name(description: str) -> str:
    description = description.upper()
    description = re.sub(r"\b(KUPO|GODOX|KNOWLED|ARRI|NSH|BLACK|SILVER|WITH|WITHOUT|W/|FOR|AND|THE)\b", " ", description)
    return "".join(re.findall(r"[A-Z0-9]+", description))[:40] or "REVIEWNAME"


def barcode(prefix: str, name: str, sequence: int | None = None) -> str:
    return f"{prefix}-{name}" if sequence is None else f"{prefix}-{name}-{sequence:04d}"


def make_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        longest = max(len(str(header)), *(len(str(row[index - 1])) for row in rows)) if rows else len(str(header))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 14), 48)


def main() -> None:
    source = SEED.read_text()
    products = sql_rows(source, "products", ["id", "name", "sku", "manufacturer_id", "tracking_mode", "daily_rate"])
    assets = sql_rows(source, "assets", ["id", "product_id", "asset_id", "status"])
    kits = sql_rows(source, "kit_instances", ["id", "kit_id", "product_id", "kit_name"])
    links = sql_rows(source, "kit_instance_components", ["id", "kit_instance_id", "asset_id"])
    stock = sql_rows(source, "stock_balances", ["product_id", "on_hand_quantity"])

    products_by_id = {row["id"]: row for row in products}
    assets_by_id = {row["id"]: row for row in assets}
    assets_by_product: dict[str, list[dict[str, str]]] = defaultdict(list)
    for asset in assets:
        product = products_by_id[asset["product_id"]]
        if product["sku"] not in CSTAND_BULK_COMPONENTS:
            assets_by_product[asset["product_id"]].append(asset)

    stock_by_product = {row["product_id"]: int(row["on_hand_quantity"]) for row in stock}
    links_by_kit: dict[str, list[dict[str, str]]] = defaultdict(list)
    for link in links:
        links_by_kit[link["kit_instance_id"]].append(link)

    kit_product_ids = {kit["product_id"] for kit in kits}
    kit_instances_by_product: dict[str, list[dict[str, str]]] = defaultdict(list)
    for kit in kits:
        kit_instances_by_product[kit["product_id"]].append(kit)

    # Build generic recipes from the provisional fixed links only to recover the
    # component *types*.  The output intentionally has no fixed kit-to-asset links.
    recipe_product_ids: dict[str, list[str]] = {}
    for parent_id, instances in kit_instances_by_product.items():
        patterns = []
        for instance in instances:
            patterns.append(tuple(sorted(assets_by_id[link["asset_id"]]["product_id"] for link in links_by_kit[instance["id"]])))
        if len(set(patterns)) != 1:
            raise ValueError(f"Inconsistent component pattern for {products_by_id[parent_id]['sku']}")
        recipe_product_ids[parent_id] = list(patterns[0])

    # Common-name payloads must remain globally unique within each prefix.
    proposed_asset_payloads: dict[str, str] = {}
    assets_by_common_name: dict[str, list[dict[str, str]]] = defaultdict(list)
    for product_id, product_assets in assets_by_product.items():
        assets_by_common_name[common_name(products_by_id[product_id])].extend(product_assets)
    for name, named_assets in assets_by_common_name.items():
        for index, asset in enumerate(sorted(named_assets, key=lambda row: row["asset_id"]), start=1):
            proposed_asset_payloads[asset["id"]] = barcode("AST", name, index)

    master_rows = []
    for parent_id, instances in sorted(kit_instances_by_product.items(), key=lambda item: products_by_id[item[0]]["sku"]):
        product = products_by_id[parent_id]
        name = KIT_COMMON_NAMES[product["sku"]]
        for index, instance in enumerate(sorted(instances, key=lambda row: row["kit_id"]), start=1):
            master_rows.append([
                product["sku"], product["name"], instance["kit_name"],
                barcode("KIT", name, index), "PROVISIONAL — client approval required",
            ])

    recipe_rows = []
    extra_quantity: Counter[str] = Counter()
    virtual_quantity_products: dict[str, tuple[str, str]] = {}
    for parent_id, component_ids in sorted(recipe_product_ids.items(), key=lambda item: products_by_id[item[0]]["sku"]):
        parent = products_by_id[parent_id]
        counts = Counter(component_ids)
        parent_name = KIT_COMMON_NAMES[parent["sku"]]
        for component_id, amount in sorted(counts.items(), key=lambda item: products_by_id[item[0]]["sku"]):
            component = products_by_id[component_id]
            if component["sku"] in CSTAND_BULK_COMPONENTS:
                model, description, short_name = CSTAND_BULK_COMPONENTS[component["sku"]]
                extra_quantity[model] += amount * len(kit_instances_by_product[parent_id])
                virtual_quantity_products[model] = (description, short_name)
                recipe_rows.append([
                    parent["sku"], parent["name"], f"KIT-{parent_name}-####",
                    model, description, short_name, "quantity", amount,
                    "Confirmed by user: shared bulk grip hardware",
                ])
                continue
            recipe_rows.append([
                parent["sku"], parent["name"], f"KIT-{parent_name}-####",
                component["sku"], component["name"], common_name(component),
                component["tracking_mode"], amount,
                "Confirmed: C-stand/product report/Godox body-only; client review still required",
            ])
        for component_sku in ARRI_BULK_COMPONENTS.get(parent["sku"], []):
            component = next(product for product in products if product["sku"] == component_sku)
            recipe_rows.append([
                parent["sku"], parent["name"], f"KIT-{parent_name}-####",
                component["sku"], component["name"], common_name(component),
                "quantity", 1, "Confirmed by user: required ARRI quantity component",
            ])

    asset_rows = []
    for product_id, product_assets in sorted(assets_by_product.items(), key=lambda item: products_by_id[item[0]]["sku"]):
        product = products_by_id[product_id]
        for asset in sorted(product_assets, key=lambda row: proposed_asset_payloads[row["id"]]):
            asset_rows.append([
                product["sku"], product["name"], common_name(product), proposed_asset_payloads[asset["id"]],
                "", "Optional — capture on physical labelling", "PROVISIONAL — client approval required",
            ])

    quantity_catalog: dict[str, dict[str, object]] = {}
    for product_id, quantity in sorted(stock_by_product.items(), key=lambda item: products_by_id[item[0]]["sku"]):
        product = products_by_id[product_id]
        quantity_catalog[product["sku"]] = {
            "description": product["name"], "common_name": common_name(product), "quantity": quantity,
        }
    for model, amount in extra_quantity.items():
        if model in quantity_catalog:
            quantity_catalog[model]["quantity"] = int(quantity_catalog[model]["quantity"]) + amount
        else:
            description, short_name = virtual_quantity_products[model]
            quantity_catalog[model] = {"description": description, "common_name": short_name, "quantity": amount}
    duplicate_names = Counter(str(record["common_name"]) for record in quantity_catalog.values())
    for record in quantity_catalog.values():
        if duplicate_names[str(record["common_name"])] > 1:
            record["common_name"] = detailed_common_name(str(record["description"]))
    quantity_rows = []
    for model, record in sorted(quantity_catalog.items()):
        name = str(record["common_name"])
        quantity_rows.append([
            model, record["description"], name, barcode("PRD", name), record["quantity"],
            "Scan opens quantity prompt", "PROVISIONAL — client approval required",
        ])
    quantity_payloads = [row[3] for row in quantity_rows]
    if len(set(quantity_payloads)) != len(quantity_payloads):
        duplicates = [value for value, count in Counter(quantity_payloads).items() if count > 1]
        raise ValueError(f"Duplicate proposed PRD barcode(s): {duplicates}")

    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read me"
    readme_rows = [
        ["Purpose", "Client-review workbook only. Do not print labels or import this dataset until approved."],
        ["Barcode status", "All KIT-, AST-, and PRD- payloads are provisional proposed values."],
        ["Kit behaviour", "Master scan chooses any available matching child assets and required quantity stock; selected children are recorded on the rental."],
        ["Manufacturer serial", "Optional field. Capture during physical labelling when a manufacturer serial exists."],
        ["Source", "Untitled spreadsheet.pdf plus confirmed user decisions; source URLs are retained in the PDF/seed, not treated as barcode authority."],
        ["Review required", "Verify item names, common-name slugs, kit recipes, and quantities before final label generation."],
    ]
    for row in readme_rows:
        readme.append(row)
    make_sheet(workbook, "Master kits", ["Parent model", "Parent description", "Kit name", "Proposed master barcode", "Status"], master_rows)
    make_sheet(workbook, "Generic kit recipes", ["Parent model", "Parent description", "Master pattern", "Child model", "Child description", "Common name", "Tracking", "Qty per kit", "Evidence / status"], recipe_rows)
    make_sheet(workbook, "Serialized assets", ["Model code", "Description", "Common name", "Proposed child barcode", "Manufacturer serial", "Serial handling", "Status"], asset_rows)
    make_sheet(workbook, "Quantity products", ["Model code", "Description", "Common name", "Proposed product barcode", "On hand", "Scan behaviour", "Status"], quantity_rows)
    make_sheet(workbook, "Validation", ["Check", "Result"], [
        ["Master kit barcodes", len(master_rows)],
        ["Serialized child barcodes", len(asset_rows)],
        ["Quantity product barcodes", len(quantity_rows)],
        ["Generic recipe lines", len(recipe_rows)],
        ["Duplicate AST payloads", "0"],
        ["Duplicate PRD payloads", "0"],
        ["Fixed kit-to-asset links", "0 in this review design"],
    ])
    for cell in readme[1]:
        cell.font = Font(bold=True)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 110

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
