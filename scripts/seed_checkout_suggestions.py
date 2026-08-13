#!/usr/bin/env python3
"""Generate SQL that imports the approved checkout-suggestion workbook.

Suggestions remain optional at rental entry. This script never creates kits,
kit components, asset hierarchy, or automatic rental lines.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "docs/generated/master_inventory_association_rules_final.xlsx"
HEADERS = (
    "association_id", "serialized_product_code", "serialized_product_name",
    "quantity_product_code", "quantity_product_name", "default_quantity",
    "quantity_rule", "recommended_v1_behavior", "reason",
)


def value(cell: object) -> str:
    return "" if cell is None else str(cell).strip()


def read_rules(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows or tuple(value(cell) for cell in rows[0]) != HEADERS:
        raise ValueError("Workbook columns do not match the approved association-rules format")

    rules: list[dict[str, object]] = []
    identifiers: set[str] = set()
    pairs: set[tuple[str, str]] = set()
    for number, row in enumerate(rows[1:], start=2):
        record = dict(zip(HEADERS, (value(cell) for cell in row), strict=True))
        if not any(record.values()):
            continue
        if record["association_id"] in identifiers:
            raise ValueError(f"Row {number}: duplicate association_id")
        pair = (record["serialized_product_code"], record["quantity_product_code"])
        if not all(pair) or not record["reason"]:
            raise ValueError(f"Row {number}: product codes and reason are required")
        if pair in pairs:
            raise ValueError(f"Row {number}: duplicate product association")
        try:
            quantity = int(record["default_quantity"])
        except ValueError as error:
            raise ValueError(f"Row {number}: default_quantity must be an integer") from error
        if quantity < 1 or record["quantity_rule"] != "FIXED_PER_ASSET" or record["recommended_v1_behavior"] != "AUTO_INCLUDE":
            raise ValueError(f"Row {number}: only approved fixed-per-asset AUTO_INCLUDE rules can be imported")
        identifiers.add(record["association_id"])
        pairs.add(pair)
        rules.append({
            "association_id": record["association_id"],
            "serialized_product_code": pair[0],
            "quantity_product_code": pair[1],
            "default_quantity": quantity,
            "reason": record["reason"],
        })
    return rules


def seed_sql(rules: list[dict[str, object]]) -> str:
    payload = json.dumps(rules, ensure_ascii=False)
    if "$suggestions$" in payload:
        raise ValueError("Workbook contains reserved SQL delimiter")
    return f"""-- Generated from master_inventory_association_rules_final.xlsx.
-- Checkout suggestions are optional at the UI; no kit/component records are created.
begin;
do $seed$
declare
  v_rule record;
begin
  for v_rule in
    select * from jsonb_to_recordset($suggestions${payload}$suggestions$::jsonb)
      as value(association_id text, serialized_product_code text, quantity_product_code text,
               default_quantity integer, reason text)
  loop
    insert into public.checkout_suggestions (
      association_id, serialized_product_id, quantity_product_id, default_quantity, reason
    )
    select v_rule.association_id, serialized_product.id, quantity_product.id,
           v_rule.default_quantity, v_rule.reason
    from public.products serialized_product
    join public.products quantity_product on quantity_product.sku = v_rule.quantity_product_code
    where serialized_product.sku = v_rule.serialized_product_code
    on conflict (association_id) do update
      set serialized_product_id = excluded.serialized_product_id,
          quantity_product_id = excluded.quantity_product_id,
          default_quantity = excluded.default_quantity,
          reason = excluded.reason;
    if not found then
      raise exception 'Suggestion % references a missing SKU', v_rule.association_id;
    end if;
  end loop;
end;
$seed$;
commit;
"""


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, help="Write SQL to this path instead of standard output")
    args = parser.parse_args()
    rules = read_rules(args.workbook)
    print(f"Validated {len(rules)} approved checkout suggestions.", file=sys.stderr)
    sql = seed_sql(rules)
    if args.out:
        args.out.write_text(sql)
    else:
        sys.stdout.write(sql)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
